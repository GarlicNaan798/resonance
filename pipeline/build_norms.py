"""
Build the psycholinguistic norm lookup tables.

This is the file that retires lib/lexicons.ts. Instead of hand-picked word
lists, every lexical value now comes from published human ratings:

  Warriner et al. (2013)  13,915 words rated for valence, arousal and dominance
                          on 1-9 scales, WITH separate means by gender (M/F),
                          age (Y/O) and education (L/H). Those demographic
                          columns are what let the audience layer be measured
                          rather than invented.

  Brysbaert et al. (2014) ~40,000 words rated for concreteness on a 1-5 scale.

Output: data/processed/norms.json
  {
    "vad":         {word: [valence, arousal, dominance]},
    "vad_demo":    {word: {"M":[v,a,d], "F":..., "Y":..., "O":..., "L":..., "H":...}},
    "concreteness":{word: value},
    "stats":       {per-dimension corpus mean/sd used for z-scoring}
  }

Values are kept on their native scales here; z-scoring happens at feature time
using the stats block, so the raw numbers stay auditable.
"""

from __future__ import annotations

import csv
import json
import os
import statistics as st

from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(ROOT, "data", "raw")
OUT_DIR = os.path.join(ROOT, "data", "processed")
OUT = os.path.join(OUT_DIR, "norms.json")

WARRINER = os.path.join(RAW, "BRM-emot-submit.csv")
CONCRETE = os.path.join(RAW, "brysbaert_concreteness.xlsx")

# Warriner column suffixes -> demographic segment key
DEMOS = {
    "M": ("V.Mean.M", "A.Mean.M", "D.Mean.M"),   # male raters
    "F": ("V.Mean.F", "A.Mean.F", "D.Mean.F"),   # female raters
    "Y": ("V.Mean.Y", "A.Mean.Y", "D.Mean.Y"),   # younger raters
    "O": ("V.Mean.O", "A.Mean.O", "D.Mean.O"),   # older raters
    "L": ("V.Mean.L", "A.Mean.L", "D.Mean.L"),   # lower education
    "H": ("V.Mean.H", "A.Mean.H", "D.Mean.H"),   # higher education
}


def _f(row: dict, key: str) -> float | None:
    raw = (row.get(key) or "").strip()
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def load_warriner() -> tuple[dict, dict]:
    vad: dict[str, list[float]] = {}
    demo: dict[str, dict[str, list[float]]] = {}

    with open(WARRINER, encoding="utf-8", newline="") as fh:
        for row in csv.DictReader(fh):
            word = (row.get("Word") or "").strip().lower()
            if not word:
                continue
            v = _f(row, "V.Mean.Sum")
            a = _f(row, "A.Mean.Sum")
            d = _f(row, "D.Mean.Sum")
            if v is None or a is None or d is None:
                continue
            vad[word] = [round(v, 3), round(a, 3), round(d, 3)]

            seg: dict[str, list[float]] = {}
            for key, (cv, ca, cd) in DEMOS.items():
                dv, da, dd = _f(row, cv), _f(row, ca), _f(row, cd)
                # keep a segment only if all three dimensions are present
                if dv is not None and da is not None and dd is not None:
                    seg[key] = [round(dv, 3), round(da, 3), round(dd, 3)]
            if seg:
                demo[word] = seg
    return vad, demo


def load_concreteness() -> dict[str, float]:
    wb = load_workbook(CONCRETE, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h else "" for h in next(rows)]
    try:
        i_word = header.index("Word")
        i_conc = header.index("Conc.M")
    except ValueError:
        raise SystemExit(f"unexpected concreteness header: {header[:8]}")

    out: dict[str, float] = {}
    for row in rows:
        if row is None or len(row) <= max(i_word, i_conc):
            continue
        w, c = row[i_word], row[i_conc]
        if w is None or c is None:
            continue
        try:
            out[str(w).strip().lower()] = round(float(c), 3)
        except (TypeError, ValueError):
            continue
    wb.close()
    return out


def main() -> None:
    # NB: plain ASCII output - the Windows console defaults to cp1252 and
    # dies on characters like the intersection sign.
    print("reading Warriner VAD...")
    vad, demo = load_warriner()
    print(f"  {len(vad):,} words with VAD; {len(demo):,} with demographic splits")

    print("reading Brysbaert concreteness...")
    conc = load_concreteness()
    print(f"  {len(conc):,} words with concreteness")

    overlap = len(set(vad) & set(conc))
    print(f"  overlap (VAD and concreteness): {overlap:,}")

    stats = {
        "valence": {"mean": st.mean(v[0] for v in vad.values()),
                    "sd": st.pstdev([v[0] for v in vad.values()])},
        "arousal": {"mean": st.mean(v[1] for v in vad.values()),
                    "sd": st.pstdev([v[1] for v in vad.values()])},
        "dominance": {"mean": st.mean(v[2] for v in vad.values()),
                      "sd": st.pstdev([v[2] for v in vad.values()])},
        "concreteness": {"mean": st.mean(conc.values()),
                         "sd": st.pstdev(list(conc.values()))},
    }
    for k, s in stats.items():
        print(f"  {k:<13} mean={s['mean']:.3f} sd={s['sd']:.3f}")

    os.makedirs(OUT_DIR, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump({"vad": vad, "vad_demo": demo,
                   "concreteness": conc, "stats": stats}, fh)
    size = os.path.getsize(OUT) / 1e6
    print(f"\nwrote {OUT} ({size:.1f} MB)")


if __name__ == "__main__":
    main()
